from flask import Flask, render_template, request, jsonify, redirect, url_for, session
from datetime import datetime, timedelta
import cv2
import os
import face_recognition
import pickle
import mysql.connector
import base64
import numpy as np
from werkzeug.security import check_password_hash
from functools import wraps
from flask import send_file 
from openpyxl import Workbook
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill
from flask_cors import CORS

# ---------------- BASE ----------------
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static")
)

CORS(app, supports_credentials=True, origins=[
    "https://facerecognition.wuaze.com",
    "https://your-app.onrender.com"
])

# Auto logout after 10 minutes of inactivity
app.permanent_session_lifetime = timedelta(minutes=10)

# IMPORTANT: change this in production
app.secret_key = os.environ.get("SECRET_KEY", "CHANGE_THIS_SECRET_KEY")

# ---------------- CONFIG ----------------
KNOWN_FACES_DIR = os.path.join(BASE_DIR, "known_faces")
ENCODINGS_FILE = os.path.join(KNOWN_FACES_DIR, "face_encodings.pkl")

DB_CONFIG = {
    "host": os.environ.get("DB_HOST"),
    "user": os.environ.get("DB_USER"),
    "password": os.environ.get("DB_PASSWORD"),
    "database": os.environ.get("DB_NAME")
}

os.makedirs(KNOWN_FACES_DIR, exist_ok=True)

cap = None

# ---------------- DB HELPER ----------------
def db():
    return mysql.connector.connect(**DB_CONFIG)

# ---------------- LOGIN REQUIRED ----------------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            session.clear()
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated   # ✅ REQUIRED

@app.before_request
def refresh_session():
    if session.get("admin_logged_in"):
        session.modified = True

# ---------------- DATABASE ----------------
def get_sections():
    conn = db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM sections ORDER BY name ASC")
    rows = cursor.fetchall()
    conn.close()
    return rows

def save_user_to_db(name, section_id):
    conn = db()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT IGNORE INTO users (name, section_id) VALUES (%s, %s)",
        (name, section_id)
    )
    conn.commit()
    conn.close()

# ---------------- ENCODINGS ----------------
def rebuild_encodings():
    known_encodings = []
    known_names = []

    for file in os.listdir(KNOWN_FACES_DIR):
        if file.lower().endswith((".jpg", ".png")):
            img = face_recognition.load_image_file(os.path.join(KNOWN_FACES_DIR, file))
            enc = face_recognition.face_encodings(img)
            if enc:
                known_encodings.append(enc[0])
                known_names.append(os.path.splitext(file)[0])

    with open(ENCODINGS_FILE, "wb") as f:
        pickle.dump({"encodings": known_encodings, "names": known_names}, f)

# ===================== LOGIN =====================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        conn = db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM admins WHERE username=%s", (username,))
        admin = cursor.fetchone()
        conn.close()

        if admin and check_password_hash(admin["password_hash"], password):
            session.clear()
            session.permanent = True
            session["admin_logged_in"] = True
            session["username"] = admin["username"]
            return redirect(url_for("index"))


        return render_template("login.html", error="Invalid username or password")

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ---------- API LOGIN ----------
@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.json
    username = data.get("username")
    password = data.get("password")

    if not username or not password:
        return jsonify({"success": False, "message": "Please fill in all fields."}), 400

    conn = db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM admins WHERE username=%s", (username,))
    admin = cursor.fetchone()
    conn.close()

    if admin and check_password_hash(admin["password_hash"], password):
        session.clear()
        session.permanent = True
        session["admin_logged_in"] = True
        session["username"] = admin["username"]
        return jsonify({"success": True, "redirect": "/"}), 200

    return jsonify({"success": False, "message": "Invalid username or password."}), 401


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"success": True, "redirect": "/login"}), 200

@app.route("/api/sections", methods=["GET"])
def api_sections():
    sections = get_sections()
    return jsonify({"sections": [{"id": s[0], "name": s[1]} for s in sections]})

# ================= FACE REGISTRATION =================
@app.route("/")
@login_required
def index():
    sections = get_sections()
    return render_template("index.html", sections=sections)


#@app.route("/start_camera", methods=["POST"])
#@login_required
#def start_camera():
    #global cap
    #cap = cv2.VideoCapture(0)
    #return jsonify({"status": "camera started"})

@app.route("/api/start_camera", methods=["POST"])
def api_start_camera():
    global cap
    cap = cv2.VideoCapture(0)
    return jsonify({"status": "camera started"})


@app.route("/capture", methods=["POST"])
@login_required
def capture():
    data = request.json
    name = data.get("name")
    section_id = data.get("section_id")
    image_data = data.get("image")

    if not name or not section_id or not image_data:
        return jsonify({"error": "Missing data"}), 400

    img_bytes = base64.b64decode(image_data.split(",")[1])
    frame = cv2.imdecode(np.frombuffer(img_bytes, np.uint8), cv2.IMREAD_COLOR)

    cv2.imwrite(os.path.join(KNOWN_FACES_DIR, f"{name}.jpg"), frame)

    save_user_to_db(name, section_id)
    rebuild_encodings()

    return jsonify({"status": "Face registered successfully"})

@app.route("/api/capture", methods=["POST"])
def api_capture():
    data = request.json
    name = data.get("name")
    section_id = data.get("section_id")
    image_data = data.get("image")

    if not name or not section_id or not image_data:
        return jsonify({"error": "Missing data"}), 400

    img_bytes = base64.b64decode(image_data.split(",")[1])
    frame = cv2.imdecode(np.frombuffer(img_bytes, np.uint8), cv2.IMREAD_COLOR)

    cv2.imwrite(os.path.join(KNOWN_FACES_DIR, f"{name}.jpg"), frame)
    save_user_to_db(name, section_id)
    rebuild_encodings()

    return jsonify({"status": "Face registered successfully"})


@app.route("/stop_camera", methods=["POST"])
@login_required
def stop_camera():
    global cap
    if cap:
        cap.release()
        cap = None
    return jsonify({"status": "camera stopped"})

@app.route("/api/stop_camera", methods=["POST"])
def api_stop_camera():
    global cap
    if cap:
        cap.release()
        cap = None
    return jsonify({"status": "camera stopped"})


# ================= DASHBOARD =================
@app.route("/dashboard")
@login_required
def dashboard():
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    page = int(request.args.get("page", 1))  # Get current page, default=1
    per_page = 5  # Limit 5 records per page
    offset = (page - 1) * per_page

    conn = db()
    cursor = conn.cursor(dictionary=True)

    # Base query
    query = """
        SELECT a.id, u.name, s.name AS section, a.date,
            a.morning_in, a.morning_out,
            a.afternoon_in, a.afternoon_out
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        LEFT JOIN sections s ON u.section_id = s.id
        WHERE 1=1
    """
    count_query = "SELECT COUNT(*) AS total FROM attendance a JOIN users u ON a.user_id = u.id LEFT JOIN sections s ON u.section_id = s.id WHERE 1=1"
    params = []

    if date_from:
        query += " AND a.date >= %s"
        count_query += " AND a.date >= %s"
        params.append(date_from)

    if date_to:
        query += " AND a.date <= %s"
        count_query += " AND a.date <= %s"
        params.append(date_to)

    query += " ORDER BY a.date DESC, u.name ASC LIMIT %s OFFSET %s"
    params_for_query = params + [per_page, offset]

    # Execute paginated query
    cursor.execute(query, params_for_query)
    records = cursor.fetchall()

    # Get total records for pagination
    cursor.execute(count_query, params)
    total_records = cursor.fetchone()["total"]
    total_pages = (total_records + per_page - 1) // per_page  # Ceiling division

    conn.close()

    # Determine status
    for r in records:
        if r["morning_in"] and r["afternoon_in"]:
            r["status"] = "Present"
        elif not r["morning_in"] and not r["afternoon_in"]:
            r["status"] = "Absent"
        else:
            r["status"] = "Partial"

    return render_template(
        "dashboard.html",
        attendance_records=records,
        current_date=datetime.now().strftime("%B %d, %Y"),
        system_status="Active",
        total_records=total_records,
        date_from=date_from,
        date_to=date_to,
        page=page,
        total_pages=total_pages
    )


@app.route("/dashboard/export")
@login_required
def export_dashboard_excel():
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    conn = db()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT
            u.name AS name,
            s.name AS section,
            a.date,
            a.morning_in,
            a.morning_out,
            a.afternoon_in,
            a.afternoon_out
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        LEFT JOIN sections s ON u.section_id = s.id
        WHERE 1=1
    """
    params = []

    if date_from:
        query += " AND a.date >= %s"
        params.append(date_from)

    if date_to:
        query += " AND a.date <= %s"
        params.append(date_to)

    query += " ORDER BY a.date DESC, u.name ASC"

    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Records"

    # --- Page setup for printing ---
    from openpyxl.worksheet.page import PageMargins

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # Landscape
    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL          # 8.5" x 13" legal size
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    ws.page_setup.fitToWidth = 1     # Fit all columns to 1 page width
    ws.page_setup.fitToHeight = 0    # Height auto
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    # ===== TITLE =====
    ws.merge_cells("A1:H1")
    ws["A1"].value = "Attendance Records"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])

    headers = [
        "Name", "Section", "Date",
        "Morning In", "Morning Out",
        "Afternoon In", "Afternoon Out", "Status"
    ]
    ws.append(headers)

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(header)
        for r in records:
            if header == "Status":
                value = (
                    "Present" if r["morning_in"] and r["afternoon_in"]
                    else "Absent" if not r["morning_in"] and not r["afternoon_in"]
                    else "Partial"
                )
            else:
                value = r[header.lower().replace(" ", "_")]
            if value:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = max_length + 5

    # Bold headers
    for col in range(1, len(headers) + 1):
        ws.cell(row=3, column=col).font = Font(bold=True)

    status_fills = {
        "Present": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Green
        "Absent": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),   # Red
        "Partial": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Yellow
    }

    start_row = 4  # header is row 3

    for idx, r in enumerate(records, start=start_row):
        status = (
            "Present" if r["morning_in"] and r["afternoon_in"]
            else "Absent" if not r["morning_in"] and not r["afternoon_in"]
            else "Partial"
        )

        row_data = [
            r["name"],
            r["section"],
            r["date"],
            r["morning_in"],
            r["morning_out"],
            r["afternoon_in"],
            r["afternoon_out"],
            status
        ]

        ws.append(row_data)

        for col in range(1, 9):
            cell = ws.cell(row=idx, column=col)

            if col in (1, 2):  # Name, Section
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if col == 8:  # Status
                cell.fill = status_fills.get(status)
                cell.font = Font(bold=True)

        ws.row_dimensions[idx].height = 28

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="attendance_dashboard.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ================= EMPLOYEES =================
@app.route("/employees")
@login_required
def employees():
    search = request.args.get("search", "")
    section_id = request.args.get("section", "")

    conn = db()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT u.id, u.name, s.name AS section, s.id AS section_id
        FROM users u
        LEFT JOIN sections s ON u.section_id = s.id
        WHERE 1=1
    """
    params = []

    if search:
        query += " AND u.name LIKE %s"
        params.append(f"%{search}%")

    if section_id:
        query += " AND u.section_id = %s"
        params.append(section_id)

    query += " ORDER BY u.name ASC"

    cursor.execute(query, params)
    employees = cursor.fetchall()

    cursor.execute("SELECT * FROM sections ORDER BY name ASC")
    sections = cursor.fetchall()

    conn.close()

    return render_template(
        "employees.html",
        employees=employees,
        sections=sections,
        total_employees=len(employees),
        search=search,
        selected_section=section_id
    )

@app.route("/employees/<int:user_id>/attendance")
@login_required
def employee_attendance(user_id):
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    conn = db()
    cursor = conn.cursor(dictionary=True)

    # Get user info
    cursor.execute("""
        SELECT u.name, s.name AS section
        FROM users u
        LEFT JOIN sections s ON u.section_id = s.id
        WHERE u.id = %s
    """, (user_id,))
    user = cursor.fetchone()

    # Base query
    query = """
        SELECT date, morning_in, morning_out, afternoon_in, afternoon_out
        FROM attendance
        WHERE user_id = %s
    """
    params = [user_id]

    # Apply date filters
    if date_from:
        query += " AND date >= %s"
        params.append(date_from)

    if date_to:
        query += " AND date <= %s"
        params.append(date_to)

    query += " ORDER BY date DESC"

    cursor.execute(query, params)
    records = cursor.fetchall()

    conn.close()

    return render_template(
        "employee_attendance.html",
        user=user,
        records=records
    )


@app.route("/employees/<int:user_id>/edit", methods=["GET", "POST"])
@login_required
def edit_employee(user_id):
    conn = db()
    cursor = conn.cursor(dictionary=True)

    if request.method == "POST":
        name = request.form["name"]
        section_id = request.form["section_id"]

        cursor.execute("""
            UPDATE users
            SET name=%s, section_id=%s
            WHERE id=%s
        """, (name, section_id or None, user_id))
        conn.commit()
        conn.close()
        return redirect("/employees")

    cursor.execute("SELECT * FROM users WHERE id=%s", (user_id,))
    employee = cursor.fetchone()

    cursor.execute("SELECT * FROM sections ORDER BY name ASC")
    sections = cursor.fetchall()

    conn.close()

    return render_template(
        "edit_employee.html",
        employee=employee,
        sections=sections
    )

@app.route("/employees/<int:user_id>/delete", methods=["POST"])
@login_required
def delete_employee(user_id):
    conn = db()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM attendance WHERE user_id=%s", (user_id,))
    if cursor.fetchone()[0] > 0:
        conn.close()
        return redirect("/employees")  # block delete if attendance exists

    cursor.execute("DELETE FROM users WHERE id=%s", (user_id,))
    conn.commit()
    conn.close()
    return redirect("/employees")

@app.route("/employees/export")
@login_required
def export_employees_excel():
    search = request.args.get("search", "")
    section_id = request.args.get("section", "")

    conn = db()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT u.name, s.name AS section
        FROM users u
        LEFT JOIN sections s ON u.section_id = s.id
        WHERE 1=1
    """
    params = []

    if search:
        query += " AND u.name LIKE %s"
        params.append(f"%{search}%")

    if section_id:
        query += " AND u.section_id = %s"
        params.append(section_id)

    query += " ORDER BY u.name ASC"
    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees List"

    # --- Page setup for printing ---
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True

    # ===== TITLE =====
    ws.merge_cells("A1:B1")
    ws["A1"].value = "Employees List"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.append([])  # blank row

    # ===== HEADERS =====
    headers = ["Name", "Section"]
    ws.append(headers)

    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill

    # ===== ROWS =====
    start_row = 4
    for idx, r in enumerate(records, start=start_row):
        ws.append([r["name"], r["section"]])
        for col in range(1, 3):
            ws.cell(row=idx, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[idx].height = 22

    # ===== AUTO-COLUMN WIDTH =====
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(header)
        for r in records:
            value = r[header.lower()]
            if value:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = max_length + 5

    # ===== SAVE AND SEND FILE =====
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="employees_list.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ---------------- RUN ----------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True) #change the port=5000 into port=8080
